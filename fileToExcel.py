import os
import pandas as pd
import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog

#####CONVERETER
def convertir_fichero(ruta_al_fichero, ruta_excel, separador):
    with open(ruta_al_fichero, 'r') as archivo:
        primera_linea = archivo.readline()
        num_columnas = len(primera_linea.split(separador))

    columnas = ['col_{}'.format(i) for i in range(1, num_columnas + 1)]
    df = pd.read_csv(ruta_al_fichero, sep=separador, names=columnas, header=None, dtype=str)

#    nombres_genericos = ['Columna {}'.format(i) for i in range(1, num_columnas + 1)]
    contador_columnas = list(range(1, num_columnas + 1))
    df = pd.concat([
        pd.DataFrame([contador_columnas], columns=columnas),
#        pd.DataFrame([nombres_genericos], columns=columnas),
        df
    ], ignore_index=True)

    with pd.ExcelWriter(ruta_excel, engine='openpyxl', mode='w', date_format='YYYYMMDD', datetime_format='YYYYMMDD') as writer:
        df.to_excel(writer, index=False, sheet_name='Datos', na_rep='')

    messagebox.showinfo("Conversión exitosa", f"El archivo Excel se ha generado en: {ruta_excel}")

def seleccionar_archivos_entrada():
    rutas_al_fichero = filedialog.askopenfilenames(title="Seleccione los ficheros de entrada",
                                                    filetypes=(("Todos los archivos", "*.*"), ("Archivos de texto", "*.txt;*.csv")))
    if rutas_al_fichero:
        separador = simpledialog.askstring("Separador", "Introduzca el carácter separador de campos:")
        if separador:
            carpeta_salida = filedialog.askdirectory(title="Seleccione la carpeta de salida")
            if carpeta_salida:
                for ruta_al_fichero in rutas_al_fichero:
                    nombre_base, _ = os.path.splitext(os.path.basename(ruta_al_fichero))
                    ruta_excel = os.path.join(carpeta_salida, f'{nombre_base}.xlsx')
                    convertir_fichero(ruta_al_fichero, ruta_excel, separador)



def seleccionar_carpeta_salida(ruta_al_fichero):
    carpeta_salida = filedialog.askdirectory(title="Seleccione la carpeta de salida")
    if carpeta_salida:
        nombre_base, _ = os.path.splitext(os.path.basename(ruta_al_fichero))
        ruta_excel = os.path.join(carpeta_salida, f'{nombre_base}.xlsx')
        return ruta_excel
    return None


####MIXER
def combinar_archivos_excel(rutas_excel, ruta_excel_salida):
    workbook_salida = openpyxl.Workbook()
    workbook_salida.remove(workbook_salida.active)  # Elimina la hoja predeterminada creada al inicializar el libro

    for ruta in rutas_excel:
        workbook_entrada = openpyxl.load_workbook(ruta, data_only=True)
        nombre_archivo = os.path.splitext(os.path.basename(ruta))[0]

        nombre_limpio = nombre_archivo.replace("EKIFD_D02_99991231_CREG", "")
        partes_nombre = nombre_limpio.split("_")
        if len(partes_nombre) > 1:
            nombre_pestaña = "_".join(partes_nombre[1:])[:30]
        else:
            nombre_pestaña = nombre_limpio[:30]

        hoja_entrada = workbook_entrada.active
        hoja_salida = workbook_salida.create_sheet(title=nombre_pestaña)
        
        for fila in hoja_entrada.iter_rows():
            hoja_salida.append([celda.value for celda in fila])

    workbook_salida.save(ruta_excel_salida)
    workbook_salida.close()
    messagebox.showinfo("Combinación exitosa", f"Los archivos Excel se han combinado en: {ruta_excel_salida}")


def seleccionar_archivos_excel():
    rutas_excel = filedialog.askopenfilenames(title="Seleccione los archivos Excel para combinar",
                                              filetypes=(("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")))
    if rutas_excel:
        ruta_excel_salida = seleccionar_archivo_excel_salida()
        if ruta_excel_salida:
            combinar_archivos_excel(rutas_excel, ruta_excel_salida)

def seleccionar_archivo_excel_salida():
    ruta_excel_salida = filedialog.asksaveasfilename(title="Guardar archivo combinado como",
                                                     filetypes=(("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")),
                                                     defaultextension=".xlsx")
    return ruta_excel_salida

#####Excel to Text-based File Converter (One file per sheet)
def convert_excel_to_text(ruta_excel, carpeta_salida, separador):
    xls = pd.read_excel(ruta_excel, sheet_name=None, dtype=str)
    
    for sheet_name, df in xls.items():
        ruta_al_fichero = os.path.join(carpeta_salida, f'{sheet_name}.txt')
        df.to_csv(ruta_al_fichero, sep=separador, index=False, header=False)
        messagebox.showinfo("Conversión exitosa", f"El archivo de texto para la pestaña '{sheet_name}' se ha generado en: {ruta_al_fichero}")

def seleccionar_archivo_excel():
    ruta_excel = filedialog.askopenfilename(title="Seleccione el archivo Excel de entrada",
                                            filetypes=(("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")))
    if ruta_excel:
        separador = simpledialog.askstring("Separador", "Introduzca el carácter separador de campos:")
        if separador:
            carpeta_salida = seleccionar_carpeta_salida()
            if carpeta_salida:
                convert_excel_to_text(ruta_excel, carpeta_salida, separador)

def seleccionar_carpeta_salida():
    carpeta_salida = filedialog.askdirectory(title="Seleccione la carpeta de salida")
    return carpeta_salida

app = tk.Tk()
app.title("Conversor/Mezclaror de ficheros a Excel")
app.geometry("800x400")

etiqueta = tk.Label(app, text="Conversor de ficheros a Excel", font=("Helvetica", 14))
etiqueta.pack(pady=20)

boton = tk.Button(app, text="Seleccionar ficheros de entrada", command=seleccionar_archivos_entrada)
boton.pack()

etiqueta = tk.Label(app, text="Mezclaror de ficheros a Excel", font=("Helvetica", 14))
etiqueta.pack(pady=20)

boton = tk.Button(app, text="Seleccionar ficheros de entrada", command=seleccionar_archivos_excel)
boton.pack()

etiqueta = tk.Label(app, text="Convertidor de Excel a ficheros de texto (un fichero por pestaña)", font=("Helvetica", 14))
etiqueta.pack(pady=20)

boton = tk.Button(app, text="Seleccionar archivo Excel de entrada", command=seleccionar_archivo_excel)
boton.pack()

app.mainloop()
